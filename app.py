from flask import Flask, render_template, Response, jsonify, request, send_file
from flask_socketio import SocketIO, emit
import cv2
import numpy as np
from ultralytics import YOLO
import time
import json
import os
from datetime import datetime
from collections import defaultdict
import uuid
import threading
import logging
import io
import csv

# ============================================
# SETUP
# ============================================
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'bird-nest-ai-2024'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# ============================================
# CONFIGURATION
# ============================================
MODEL_PATH = 'best (1).pt'
DETECTION_THRESHOLD = 0.456  # Ngưỡng phát hiện
HARVESTING_THRESHOLD = 0.6   # Ngưỡng thu hoạch (60%)
IOU_THRESHOLD = 0.4          # Ngưỡng gom nhóm tổ

# Số lần detect >= 60% để tính là 1 tổ thu hoạch
MIN_HARVEST_COUNT = 3

SESSIONS_DIR = 'sessions'
os.makedirs(SESSIONS_DIR, exist_ok=True)

# ============================================
# CLASS NAMES
# ============================================
NEST_CLASS = 'to_yen'
EGG_CLASS = 'trung_chim'
CHICK_CLASS = 'chim_non'
ADULT_CLASS = 'chim_truong_thanh'

# ============================================
# GLOBAL STATE
# ============================================
class SystemState:
    IDLE = "IDLE"
    RUNNING = "RUNNING"
    STOPPING = "STOPPING"
    PROCESSING = "PROCESSING"
    RESULT_READY = "RESULT_READY"
    ERROR = "ERROR"

class HarvestStatus:
    READY = "ready"
    WARNING = "warning"
    NOT_READY = "not_ready"

current_state = SystemState.IDLE
current_session = None
model = None
camera = None
available_cameras = []

# ============================================
# SESSION CLASS
# ============================================
class Session:
    def __init__(self, session_id, camera_id):
        self.id = session_id
        self.camera_id = camera_id
        self.start_time = datetime.now()
        self.end_time = None
        self.duration = 0
        self.all_detections = []
        self.nests = {}
        self.results = {}
        self.state = SystemState.IDLE
        
        # Realtime tracking
        self.nest_tracker = {}  # {nest_key: {bbox, high_conf_count, low_conf_count, ...}}
        self.unique_nests_count = 0
    
    def to_dict(self):
        return {
            'id': self.id,
            'camera_id': self.camera_id,
            'start_time': self.start_time.isoformat(),
            'end_time': self.end_time.isoformat() if self.end_time else None,
            'duration': self.duration,
            'state': self.state,
            'total_detections': len(self.all_detections),
            'total_nests': len(self.nests),
            'unique_nests': self.unique_nests_count,
            'results': self.results
        }
    
    def save(self):
        filepath = os.path.join(SESSIONS_DIR, f"{self.id}.json")
        data = self.to_dict()
        data['nests'] = list(self.nests.values())
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        logger.info(f"💾 Saved: {filepath}")

# ============================================
# HELPER FUNCTIONS
# ============================================
def calculate_iou(box1, box2):
    x1 = max(box1[0], box2[0])
    y1 = max(box1[1], box2[1])
    x2 = min(box1[2], box2[2])
    y2 = min(box1[3], box2[3])
    
    if x2 < x1 or y2 < y1:
        return 0.0
    
    inter_area = (x2 - x1) * (y2 - y1)
    box1_area = (box1[2] - box1[0]) * (box1[3] - box1[1])
    box2_area = (box2[2] - box2[0]) * (box2[3] - box2[1])
    union_area = box1_area + box2_area - inter_area
    
    return inter_area / union_area if union_area > 0 else 0

def find_matching_nest(bbox, tracker_dict, threshold=0.3):
    """Tìm nest có IoU cao nhất"""
    best_key = None
    best_iou = 0
    for key, data in tracker_dict.items():
        iou = calculate_iou(bbox, data['bbox'])
        if iou > best_iou and iou > threshold:
            best_iou = iou
            best_key = key
    return best_key

# ============================================
# LOAD MODEL & CAMERA
# ============================================
def load_yolo_model():
    global model
    try:
        if os.path.exists(MODEL_PATH):
            model = YOLO(MODEL_PATH)
            logger.info(f"✅ Model loaded: {model.names}")
            return True
        logger.error(f"❌ Model not found: {MODEL_PATH}")
        return False
    except Exception as e:
        logger.error(f"❌ Model error: {e}")
        return False

def detect_cameras():
    global available_cameras
    available_cameras = []
    
    for i in range(5):
        for backend in [cv2.CAP_DSHOW, cv2.CAP_ANY]:
            try:
                cap = cv2.VideoCapture(i, backend)
                if cap.isOpened():
                    ret, frame = cap.read()
                    if ret:
                        h, w = frame.shape[:2]
                        available_cameras.append({
                            'id': i, 'name': f'Camera {i}',
                            'resolution': f'{w}x{h}', 'backend': backend
                        })
                        cap.release()
                        break
                    cap.release()
            except:
                pass
    
    logger.info(f"Found {len(available_cameras)} camera(s)")
    return available_cameras

def init_camera(camera_id=0):
    global camera
    if camera:
        try: camera.release()
        except: pass
    
    for backend in [cv2.CAP_DSHOW, cv2.CAP_ANY]:
        try:
            camera = cv2.VideoCapture(camera_id, backend)
            if camera.isOpened():
                camera.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
                camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
                ret, _ = camera.read()
                if ret:
                    logger.info(f"✅ Camera {camera_id} ready")
                    return True
                camera.release()
        except:
            pass
    return False

# ============================================
# VIDEO STREAM + DETECTION
# ============================================
frame_counter = 0

def generate_frames():
    global frame_counter, current_session
    
    while True:
        if camera is None or not camera.isOpened():
            dummy = np.zeros((480, 640, 3), dtype=np.uint8)
            cv2.putText(dummy, "NO CAMERA", (200, 240), cv2.FONT_HERSHEY_SIMPLEX, 1.5, (0,0,255), 3)
            ret, buffer = cv2.imencode('.jpg', dummy)
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
            time.sleep(0.1)
            continue
        
        success, frame = camera.read()
        if not success:
            time.sleep(0.05)
            continue
        
        if current_state == SystemState.RUNNING and model and current_session:
            try:
                results = model(frame, conf=DETECTION_THRESHOLD, verbose=False)
                current_time = time.time()
                
                for result in results:
                    for box in result.boxes:
                        x1, y1, x2, y2 = map(int, box.xyxy[0])
                        conf = float(box.conf[0])
                        cls = int(box.cls[0])
                        class_name = model.names[cls]
                        bbox = [x1, y1, x2, y2]
                        
                        # Lưu detection
                        current_session.all_detections.append({
                            'frame_id': frame_counter,
                            'timestamp': current_time,
                            'class': class_name,
                            'bbox': bbox,
                            'confidence': conf
                        })
                        
                        # Realtime tracking cho to_yen
                        if class_name == NEST_CLASS:
                            matching_key = find_matching_nest(bbox, current_session.nest_tracker)
                            
                            if matching_key is None:
                                # Tổ mới
                                new_key = f"nest_{len(current_session.nest_tracker)}"
                                current_session.nest_tracker[new_key] = {
                                    'bbox': bbox,
                                    'high_conf_count': 1 if conf >= HARVESTING_THRESHOLD else 0,
                                    'low_conf_count': 1 if conf < HARVESTING_THRESHOLD else 0,
                                    'max_conf': conf,
                                    'total_count': 1
                                }
                            else:
                                # Cập nhật tổ đã có
                                tracker = current_session.nest_tracker[matching_key]
                                if conf >= HARVESTING_THRESHOLD:
                                    tracker['high_conf_count'] += 1
                                else:
                                    tracker['low_conf_count'] += 1
                                tracker['total_count'] += 1
                                tracker['max_conf'] = max(tracker['max_conf'], conf)
                                # Update bbox
                                tracker['bbox'] = [
                                    int(tracker['bbox'][i] * 0.8 + bbox[i] * 0.2)
                                    for i in range(4)
                                ]
                            
                            # Cập nhật số tổ duy nhất (có >= 3 lần detect)
                            current_session.unique_nests_count = sum(
                                1 for t in current_session.nest_tracker.values()
                                if t['total_count'] >= 3
                            )
                        
                        # Vẽ box
                        if class_name == NEST_CLASS:
                            color = (0, 255, 0) if conf >= HARVESTING_THRESHOLD else (0, 255, 255)
                            label = f"To yen {conf:.0%}"
                        elif class_name == EGG_CLASS:
                            color = (0, 200, 255)
                            label = f"Trung {conf:.0%}"
                        elif class_name == CHICK_CLASS:
                            color = (0, 165, 255)
                            label = f"Chim non {conf:.0%}"
                        elif class_name == ADULT_CLASS:
                            color = (0, 0, 255)
                            label = f"Chim TT {conf:.0%}"
                        else:
                            color = (255, 255, 255)
                            label = f"{class_name} {conf:.0%}"
                        
                        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                        cv2.putText(frame, label, (x1, y1 - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
                        
                        socketio.emit('detection', {
                            'class': class_name, 
                            'confidence': conf,
                            'total_detections': len(current_session.all_detections),
                            'unique_nests': current_session.unique_nests_count
                        })
                
                frame_counter += 1
                
            except Exception as e:
                logger.error(f"Detection error: {e}")
        
        # Overlay
        if current_state == SystemState.RUNNING and current_session:
            elapsed = int(time.time() - current_session.start_time.timestamp())
            mins, secs = divmod(elapsed, 60)
            cv2.circle(frame, (25, 30), 8, (0, 0, 255), -1)
            cv2.putText(frame, f"REC {mins:02d}:{secs:02d}", (40, 38), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255,255,255), 2)
            cv2.putText(frame, f"Detections: {len(current_session.all_detections)}", (20, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,0), 2)
            cv2.putText(frame, f"Unique nests: {current_session.unique_nests_count}", (20, 95), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,255), 2)
        
        ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')

# ============================================
# PROCESSING
# ============================================
def process_session(session):
    """Xử lý kết quả sau khi dừng"""
    global current_state
    
    try:
        current_state = SystemState.PROCESSING
        socketio.emit('state_change', {'state': current_state})
        
        logger.info(f"\n📊 Processing session {session.id}")
        logger.info(f"   Total detections: {len(session.all_detections)}")
        logger.info(f"   Tracked nests: {len(session.nest_tracker)}")
        
        # Phân loại từng tổ
        socketio.emit('processing_step', {'step': 1, 'message': 'Phân loại tổ...'})
        
        for idx, (key, tracker) in enumerate(session.nest_tracker.items()):
            nest_id = f"TO-{idx+1:03d}"
            
            high_count = tracker['high_conf_count']  # Số lần >= 60%
            low_count = tracker['low_conf_count']    # Số lần < 60%
            total_count = tracker['total_count']
            max_conf = tracker['max_conf']
            
            # LOGIC MỚI:
            # - Có >= 3 lần detect với conf >= 60% -> SẴN SÀNG THU HOẠCH
            # - Có >= 3 lần detect nhưng conf < 60% -> CẦN KIỂM TRA
            # - < 3 lần detect -> không tính
            
            if total_count < 3:
                # Không đủ 3 lần detect -> bỏ qua
                continue
            
            if high_count >= MIN_HARVEST_COUNT:
                status = HarvestStatus.READY
                reason = f"Sẵn sàng thu hoạch ({high_count} lần >= 60%)"
                harvest_safe = True
            elif high_count > 0:
                status = HarvestStatus.WARNING
                reason = f"Cần kiểm tra ({high_count} lần >= 60%, {low_count} lần < 60%)"
                harvest_safe = False
            else:
                status = HarvestStatus.WARNING
                reason = f"Cần kiểm tra (tất cả {total_count} lần < 60%)"
                harvest_safe = False
            
            session.nests[nest_id] = {
                'id': nest_id,
                'bbox': [float(x) for x in tracker['bbox']],
                'status': status,
                'reason': reason,
                'harvest_safe': harvest_safe,
                'high_conf_count': high_count,
                'low_conf_count': low_count,
                'total_count': total_count,
                'max_confidence': float(max_conf),
                'has_egg': False,
                'has_chick': False,
                'has_adult': False
            }
            
            logger.info(f"   {nest_id}: {status} - {reason}")
        
        # Tính KPI
        socketio.emit('processing_step', {'step': 2, 'message': 'Tính toán KPI...'})
        
        total = len(session.nests)
        ready = sum(1 for n in session.nests.values() if n['status'] == HarvestStatus.READY)
        warning = sum(1 for n in session.nests.values() if n['status'] == HarvestStatus.WARNING)
        
        harvest_rate = (ready / total * 100) if total > 0 else 0
        
        session.results = {
            'total_nests': total,
            'ready_nests': ready,
            'warning_nests': warning,
            'not_ready_nests': 0,
            'harvest_rate': round(harvest_rate, 2),
            'breakdown': {
                'ready': ready,
                'warning': warning,
                'not_ready': 0,
                'has_egg': 0,
                'has_chick': 0,
                'has_adult': 0,
            },
            'total_detections': len(session.all_detections)
        }
        
        logger.info(f"\n✅ KẾT QUẢ:")
        logger.info(f"   Tổng số tổ: {total}")
        logger.info(f"   Sẵn sàng: {ready}")
        logger.info(f"   Kiểm tra: {warning}")
        logger.info(f"   Tỷ lệ: {harvest_rate:.1f}%\n")
        
        session.save()
        
        current_state = SystemState.RESULT_READY
        socketio.emit('state_change', {'state': current_state})
        socketio.emit('processing_complete', {'results': session.results})
        socketio.emit('results', {
            'session': session.to_dict(),
            'nests': list(session.nests.values())
        })
        
    except Exception as e:
        logger.error(f"❌ Processing error: {e}")
        import traceback
        traceback.print_exc()
        current_state = SystemState.ERROR
        socketio.emit('state_change', {'state': current_state, 'error': str(e)})

# ============================================
# ROUTES
# ============================================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/system/status')
def system_status():
    return jsonify({
        'state': current_state,
        'model_loaded': model is not None,
        'camera_connected': camera is not None and camera.isOpened(),
        'cameras': available_cameras,
        'session': current_session.to_dict() if current_session else None,
        'config': {
            'detection_threshold': DETECTION_THRESHOLD,
            'harvesting_threshold': HARVESTING_THRESHOLD,
            'min_harvest_count': MIN_HARVEST_COUNT
        }
    })

@app.route('/api/cameras')
def get_cameras():
    return jsonify({'cameras': available_cameras})

@app.route('/api/sessions/<session_id>')
def get_session(session_id):
    filepath = os.path.join(SESSIONS_DIR, f"{session_id}.json")
    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            return jsonify(json.load(f))
    return jsonify({'error': 'Not found'}), 404

# ============================================
# EXPORT ROUTES
# ============================================
@app.route('/api/export/json/<session_id>')
def export_json(session_id):
    try:
        if current_session and current_session.id == session_id:
            data = {'session': current_session.to_dict(), 'nests': list(current_session.nests.values())}
        else:
            filepath = os.path.join(SESSIONS_DIR, f"{session_id}.json")
            if not os.path.exists(filepath):
                return jsonify({'error': 'Not found'}), 404
            with open(filepath, 'r') as f:
                data = json.load(f)
        
        output = io.BytesIO()
        output.write(json.dumps(data, indent=2, ensure_ascii=False).encode('utf-8'))
        output.seek(0)
        
        return send_file(output, mimetype='application/json', as_attachment=True,
                        download_name=f"bao_cao_{session_id}.json")
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/csv/<session_id>')
def export_csv(session_id):
    try:
        if current_session and current_session.id == session_id:
            session_data = current_session.to_dict()
            nests = list(current_session.nests.values())
        else:
            filepath = os.path.join(SESSIONS_DIR, f"{session_id}.json")
            if not os.path.exists(filepath):
                return jsonify({'error': 'Not found'}), 404
            with open(filepath, 'r') as f:
                session_data = json.load(f)
                nests = session_data.get('nests', [])
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['BAO CAO PHAN TICH TO YEN'])
        writer.writerow(['Session ID', session_data.get('id', '')])
        writer.writerow([])
        
        results = session_data.get('results', {})
        writer.writerow(['Tong so to', results.get('total_nests', 0)])
        writer.writerow(['San sang thu hoach', results.get('ready_nests', 0)])
        writer.writerow(['Can kiem tra', results.get('warning_nests', 0)])
        writer.writerow(['Ty le thu hoach', f"{results.get('harvest_rate', 0)}%"])
        writer.writerow([])
        
        writer.writerow(['ID', 'Trang thai', 'So lan >=60%', 'So lan <60%', 'Tong', 'Ghi chu'])
        for nest in nests:
            status_vn = {'ready': 'San sang', 'warning': 'Kiem tra'}.get(nest.get('status'), nest.get('status'))
            writer.writerow([
                nest.get('id'),
                status_vn,
                nest.get('high_conf_count', 0),
                nest.get('low_conf_count', 0),
                nest.get('total_count', 0),
                nest.get('reason', '')
            ])
        
        output.seek(0)
        bytes_out = io.BytesIO()
        bytes_out.write(output.getvalue().encode('utf-8-sig'))
        bytes_out.seek(0)
        
        return send_file(bytes_out, mimetype='text/csv', as_attachment=True,
                        download_name=f"bao_cao_{session_id}.csv")
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/excel/<session_id>')
def export_excel(session_id):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'Cần cài openpyxl: pip install openpyxl'}), 500
    
    try:
        if current_session and current_session.id == session_id:
            session_data = current_session.to_dict()
            nests = list(current_session.nests.values())
        else:
            filepath = os.path.join(SESSIONS_DIR, f"{session_id}.json")
            if not os.path.exists(filepath):
                return jsonify({'error': 'Not found'}), 404
            with open(filepath, 'r') as f:
                session_data = json.load(f)
                nests = session_data.get('nests', [])
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bao Cao"
        
        # Styles
        header_fill = PatternFill(start_color="2D5016", end_color="2D5016", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        ready_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        warning_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        
        # Title
        ws['A1'] = "BAO CAO PHAN TICH TO YEN"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Summary
        results = session_data.get('results', {})
        ws['A3'] = "Tong so to:"
        ws['B3'] = results.get('total_nests', 0)
        ws['A4'] = "San sang thu hoach:"
        ws['B4'] = results.get('ready_nests', 0)
        ws['A5'] = "Can kiem tra:"
        ws['B5'] = results.get('warning_nests', 0)
        ws['A6'] = "Ty le thu hoach:"
        ws['B6'] = f"{results.get('harvest_rate', 0)}%"
        
        # Headers
        headers = ['ID', 'Trang thai', 'So lan >=60%', 'So lan <60%', 'Tong', 'Max Conf', 'Ghi chu']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for row, nest in enumerate(nests, 9):
            status = nest.get('status', '')
            status_vn = {'ready': 'San sang', 'warning': 'Kiem tra'}.get(status, status)
            
            ws.cell(row=row, column=1, value=nest.get('id'))
            ws.cell(row=row, column=2, value=status_vn)
            ws.cell(row=row, column=3, value=nest.get('high_conf_count', 0))
            ws.cell(row=row, column=4, value=nest.get('low_conf_count', 0))
            ws.cell(row=row, column=5, value=nest.get('total_count', 0))
            ws.cell(row=row, column=6, value=f"{nest.get('max_confidence', 0)*100:.1f}%")
            ws.cell(row=row, column=7, value=nest.get('reason', ''))
            
            fill = ready_fill if status == 'ready' else warning_fill
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = fill
        
        # Column widths
        for col, w in enumerate([10, 12, 14, 14, 10, 12, 45], 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f"bao_cao_{session_id}.xlsx")
    except Exception as e:
        logger.error(f"Excel error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/pdf/<session_id>')
def export_pdf(session_id):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return jsonify({'error': 'Cần cài reportlab: pip install reportlab'}), 500
    
    try:
        if current_session and current_session.id == session_id:
            session_data = current_session.to_dict()
            nests = list(current_session.nests.values())
        else:
            filepath = os.path.join(SESSIONS_DIR, f"{session_id}.json")
            if not os.path.exists(filepath):
                return jsonify({'error': 'Not found'}), 404
            with open(filepath, 'r') as f:
                session_data = json.load(f)
                nests = session_data.get('nests', [])
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        elements.append(Paragraph("BAO CAO PHAN TICH TO YEN", styles['Heading1']))
        elements.append(Spacer(1, 20))
        
        # Summary
        results = session_data.get('results', {})
        summary_data = [
            ['Chi tieu', 'Gia tri'],
            ['Tong so to', str(results.get('total_nests', 0))],
            ['San sang thu hoach', str(results.get('ready_nests', 0))],
            ['Can kiem tra', str(results.get('warning_nests', 0))],
            ['Ty le thu hoach', f"{results.get('harvest_rate', 0)}%"],
        ]
        
        summary_table = Table(summary_data, colWidths=[200, 100])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.18, 0.31, 0.09)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # Detail
        elements.append(Paragraph("Chi tiet tung to", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        detail_data = [['ID', 'Trang thai', '>=60%', '<60%', 'Max Conf']]
        for nest in nests:
            status_vn = {'ready': 'San sang', 'warning': 'Kiem tra'}.get(nest.get('status'), '')
            detail_data.append([
                nest.get('id'),
                status_vn,
                str(nest.get('high_conf_count', 0)),
                str(nest.get('low_conf_count', 0)),
                f"{nest.get('max_confidence', 0)*100:.1f}%"
            ])
        
        detail_table = Table(detail_data, colWidths=[60, 80, 60, 60, 70])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.18, 0.31, 0.09)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(detail_table)
        
        doc.build(elements)
        output.seek(0)
        
        return send_file(output, mimetype='application/pdf', as_attachment=True,
                        download_name=f"bao_cao_{session_id}.pdf")
    except Exception as e:
        logger.error(f"PDF error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/formats')
def export_formats():
    # Check what's available
    excel_ok = False
    pdf_ok = False
    try:
        import openpyxl
        excel_ok = True
    except: pass
    try:
        import reportlab
        pdf_ok = True
    except: pass
    
    return jsonify({
        'formats': [
            {'id': 'json', 'name': 'JSON', 'available': True},
            {'id': 'csv', 'name': 'CSV', 'available': True},
            {'id': 'excel', 'name': 'Excel', 'available': excel_ok},
            {'id': 'pdf', 'name': 'PDF', 'available': pdf_ok},
        ]
    })

# ============================================
# SOCKET HANDLERS
# ============================================
@socketio.on('start_session')
def handle_start_session(data):
    global current_state, current_session, frame_counter
    
    camera_id = data.get('camera_id', 0)
    if not init_camera(camera_id):
        emit('error', {'message': 'Camera failed'})
        return
    
    session_id = str(uuid.uuid4())[:8]
    current_session = Session(session_id, camera_id)
    current_session.state = SystemState.RUNNING
    
    current_state = SystemState.RUNNING
    frame_counter = 0
    
    logger.info(f"\n🚀 Session started: {session_id}\n")
    
    emit('session_started', {'session_id': session_id})
    emit('state_change', {'state': current_state})

@socketio.on('stop_session')
def handle_stop_session():
    global current_state, current_session
    
    if current_session:
        current_session.end_time = datetime.now()
        current_session.duration = int((current_session.end_time - current_session.start_time).total_seconds())
        current_session.state = SystemState.STOPPING
        
        current_state = SystemState.STOPPING
        emit('state_change', {'state': current_state})
        
        logger.info(f"\n⏹️ Session stopped: {current_session.id}")
        logger.info(f"   Duration: {current_session.duration}s")
        logger.info(f"   Detections: {len(current_session.all_detections)}")
        
        threading.Thread(target=process_session, args=(current_session,), daemon=True).start()

@socketio.on('get_results')
def handle_get_results():
    if current_session and current_state == SystemState.RESULT_READY:
        emit('results', {
            'session': current_session.to_dict(),
            'nests': list(current_session.nests.values())
        })

# ============================================
# MAIN
# ============================================
if __name__ == '__main__':
    print("\n" + "=" * 55)
    print("🪺 HỆ THỐNG PHÂN TÍCH TỔ YẾN v4.0")
    print("=" * 55)
    print(f"\n📊 NGƯỠNG:")
    print(f"   • Detection: {DETECTION_THRESHOLD:.1%}")
    print(f"   • Harvesting: {HARVESTING_THRESHOLD:.0%}")
    print(f"   • Min count: {MIN_HARVEST_COUNT} lần")
    print(f"\n📋 LOGIC:")
    print(f"   ✅ Sẵn sàng: >= {MIN_HARVEST_COUNT} lần detect với conf >= 60%")
    print(f"   ⚠️ Kiểm tra: < {MIN_HARVEST_COUNT} lần >= 60%")
    print("=" * 55)
    
    load_yolo_model()
    detect_cameras()
    if available_cameras:
        init_camera(available_cameras[0]['id'])
    
    print(f"\n🌐 http://localhost:5000")
    print("=" * 55 + "\n")
    
    socketio.run(app, host='0.0.0.0', port=5000, debug=False, allow_unsafe_werkzeug=True)